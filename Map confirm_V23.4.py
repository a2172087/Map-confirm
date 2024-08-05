import os
import re
import sys
import traceback
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QVBoxLayout, QPushButton, QFileDialog, QLineEdit, QMessageBox, QCheckBox, QComboBox, QInputDialog, QDialog, QDialogButtonBox
from PyQt5.QtGui import QPixmap, QFont, QImage, QPainter, QColor, QIcon
from PyQt5.QtCore import Qt, QRect
import qtmodern.styles
import qtmodern.windows
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import socket
from datetime import datetime
from collections import OrderedDict
import shutil
import py7zr
import tempfile
import datetime

if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

icon_path = os.path.join(application_path, 'format.ico')

class ImageClassifier(QWidget):
    def __init__(self):
        super().__init__()

        # 呼叫 check_version 方法
        self.check_version()
        self.save_log()

        # 設定視窗標題和圖示
        self.setWindowTitle('Map Confirm')
        app.setWindowIcon(QIcon(icon_path))
        self.resize(800, 600)
        
        # 建立導入MAP圖按鈕
        self.importBtn = QPushButton('導入Defect Map picture', self)
        self.importBtn.clicked.connect(self.importMap)
        
        # 建立選擇確認照片資料夾按鈕
        self.confirmPathBtn = QPushButton('選擇貨批的Defcet資料夾路徑', self)
        self.confirmPathBtn.clicked.connect(self.selectConfirmPath)
        self.confirmPathBtn.setEnabled(False)
        
        # 建立執行按鈕
        self.runSaveBtn = QPushButton('執行', self)
        self.runSaveBtn.clicked.connect(self.executeSave)

        # 建立輸出Defect die座標按鈕
        self.outputDefectDieBtn = QPushButton('輸出Defect die座標', self)
        self.outputDefectDieBtn.clicked.connect(self.outputDefectDieCoordinates)
        self.outputDefectDieBtn.hide()

        # 建立不要覆蓋已存在的Issue code的複選框
        self.overwriteCheckBox = QCheckBox('不要覆蓋已存在的Issue code', self)
        self.overwriteCheckBox.setChecked(True)  # 預設打勾
        self.overwriteCheckBox.hide()

        self.initializeAllMapsBtn = QPushButton('初始化所有MAP', self)
        self.initializeAllMapsBtn.clicked.connect(self.initializeAllMaps)
        self.initializeAllMapsBtn.hide()

        # 建立Defect Chip summary按鈕
        self.defectChipSummaryBtn = QPushButton('Defect Chip summary', self)
        self.defectChipSummaryBtn.clicked.connect(self.performDefectChipSummary)
        self.defectChipSummaryBtn.hide()

        # 建立輸出Yield loss result按鈕
        self.outputYieldLossResultBtn = QPushButton('輸出Yield loss result', self)
        self.outputYieldLossResultBtn.clicked.connect(self.outputYieldLossResult)
        self.outputYieldLossResultBtn.hide()
        
        # 建立MAP檔案路徑標籤
        self.pathLabel = QLineEdit(self)
        self.pathLabel.hide()
        
        # 建立確認照片資料夾路徑標籤
        self.confirmPathLabel = QLineEdit(self)
        self.confirmPathLabel.hide()

        # 建立選擇Sheet的下拉選單
        self.sheetComboBox = QComboBox(self)
        self.sheetComboBox.currentIndexChanged.connect(self.displaySelectedSheet)
        self.sheetComboBox.hide()
        
        # 建立圖片標籤
        self.imageLabel = QLabel(self)
        
        # 設定預設的填充顏色
        self.start_color = "bc8f8f"
        self.end_color = "bc8f8f"
        
        # 建立佈局並將元件加入佈局
        layout = QVBoxLayout()
        layout.addWidget(self.importBtn)
        layout.addWidget(self.pathLabel)
        layout.addWidget(self.confirmPathBtn)
        layout.addWidget(self.confirmPathLabel)
        layout.addWidget(self.overwriteCheckBox)
        layout.addWidget(self.sheetComboBox)
        layout.addWidget(self.initializeAllMapsBtn)
        layout.addWidget(self.imageLabel)
        layout.addWidget(self.runSaveBtn)
        layout.addWidget(self.outputDefectDieBtn)
        layout.addWidget(self.outputYieldLossResultBtn)
        layout.addWidget(self.defectChipSummaryBtn)
        self.setLayout(layout)
        
        # 檢查是否顯示執行按鈕
        self.checkShowRunButton()

        # 檢查版本是否為最新版本
        self.check_version()

        # 定義 Issue Code 選擇對話框的相關屬性和方法
        self.issue_code_dialog = None
        self.selected_issue_codes = []

        # 定義每個子資料夾對應的顏色
        self.folder_color_mapping = OrderedDict([
            ('102_Foreign_material(09)', 'FFB6C1'),  # 淺粉紅色 
            ('000_Particle(16)', 'DA70D6'),  # 淺紫色
            ('200_Probe_Mark_Shift(10)', 'FFDAB9'),  # 淺橙色
            ('502_Bump foreign Material(25)', 'F0E68C'),  # 卡其色
            ('100_Process_Defect(07)', '90EE90'),  # 淺綠色
            ('205_Al_particle_out_of_pad(0F)', 'DEB887'),  # 實木色  
            ('999_Al_particle(18)', 'CD5C5C'),  # 印度紅
            ('101_Wafer Scratch(08)', 'FFC0CB'),  # 淺紅色
            ('202_PM area out spec.(12)', 'ADFF2F'),  # 綠黃色
            ('500_Bump PM shift(23)', 'DDA0DD'),  # 淺紫色
            ('501_Bump scratch(24)', 'DA70D6'),  # 淺紫色
            ('186_Other(BA)', 'FFDAB9'),  # 核桃色
            ('151_Ugly_Die(2D)', 'FFFF00'),  # 淺黃色
            ('507_Bump PM diameter out of spec(0E)', '40E0D0'), # 綠松石色
            ('201_PM No. Out Spec(11)', '#FF7F50'),  # 珊瑚色
            ('117_Pad discoloration(1B)', '#6A5ACD'),  # 板岩藍
            ('505_Irregular bump(0A)', '#FFD700'),  # 金色
            ('203_Probing Void(13)', '#3CB371'),  # 中海綠色
            ('204_Missing Probe Mark(14)', '#DC143C'),  # 深紅色
            ('100_Surface(Incoming defect)(1C)', '#FF8C00'),  # 深橙色
            ('503_Missing bump(26)', '#BA55D3'),  # 中蘭花紫
            ('504_Bump residue(27)', '#4682B4'),  # 鋼青色
            ('506_Bump house defect(28)', '#A0522D'),  # 赭色     
            ('115_Large defect(31)', '#98FF98'),  #薄荷綠       
            ('510_Large bump(32)', '#FFE5B4'),  #淺杏色
            ('521_small bump(3D)', '#FFFACD'),   #淺檸檬黃         
            ('522_380 special PM shift(3E)', '#F08080'),  #淺珊瑚色        
        ])

        self.No_folder_color_mapping = {
            '103': '103_AVI Defect',
            '105': '105_Die Screen Defect',
            '109': '109_FAB IQA Defect',
            '126': '126_Die crack',
            '127': '127_Wafer surface scratch',
            '146': '146_Bump defect',
            '507': '507_Bump P/M diameter out of spec.',
            '681': '681_313_GPAT', #Customer No.313
        }

        # 定義子資料夾的優先度,數字越小優先度越高
        self.subfolder_priority = {
            '151_Ugly_Die(2D)': 1,
            '203_Probing Void(13)':2, #3
            '117_Pad discoloration(1B)':3, #4
            '200_Probe_Mark_Shift(10)': 4,  #5
            '522_380 special PM shift(3E)':5, #5
            '507_Bump PM diameter out of spec(0E)' :6, #6
            '500_Bump PM shift(23)': 7, #7
            '202_PM area out spec.(12)':8, #8
            '204_Missing Probe Mark(14)':9,   #9
            '101_Wafer Scratch(08)': 10, #10
            '115_Large defect(31)' :11, #11
            '102_Foreign_material(09)': 12,  #12
            '201_PM No. Out Spec(11)' :13, #13
            '205_Al_particle_out_of_pad(0F)': 14, #14
            '502_Bump foreign Material(25)': 15, #15
            '501_Bump scratch(24)': 16, #16
            '100_Process_Defect(07)': 17, #17
            '503_Missing bump(26)' :18, #18
            '506_Bump house defect(28)' :19, #19
            '504_Bump residue(27)' :20, #20
            '510_Large bump(32)' :21, #21
            '505_Irregular bump(0A)' :22, #22
            '521_small bump(3D)' :23, #23
            '000_Particle(16)': 24, #98
            '100_Surface(Incoming defect)(1C)' :25,
            '999_Al_particle(18)': 26,
            '186_Other(BA)': 27,    
        }
        
    def importMap(self):
        # 開啟檔案對話框選擇MAP檔案
        path, _ = QFileDialog.getOpenFileName(self, '選擇MAP檔案', filter="Excel files (*.xlsx)")
        if path:
            self.pathLabel.setText(path)
            self.confirmPathBtn.setEnabled(True)
            self.pathLabel.show()
            self.outputDefectDieBtn.show()
            self.initializeAllMapsBtn.show()
            self.confirmPathBtn.show()
            self.outputYieldLossResultBtn.show()
            self.defectChipSummaryBtn.show()
            self.defectChipSummaryBtn.setEnabled(False)
            self.sheetComboBox.show()
            self.overwriteCheckBox.show()

            # 載入選擇的Excel檔案
            workbook = openpyxl.load_workbook(path)
            
            # 清空下拉選單
            self.sheetComboBox.clear()
            
            # 將所有Sheet名稱加入下拉選單
            for sheet_name in workbook.sheetnames:
                self.sheetComboBox.addItem(sheet_name)
            
            # 顯示第一個Sheet的MAP圖
            self.displaySelectedSheet()

            self.checkShowRunButton()

    def initializeAllMaps(self):
        excel_path = self.pathLabel.text()
        if excel_path:
            workbook = openpyxl.load_workbook(excel_path)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        fill_color = self.get_fill_color(cell)
                        if fill_color and (fill_color.red(), fill_color.green(), fill_color.blue()) != (198, 226, 255):
                            cell.fill = PatternFill(start_color='C6E2FF', end_color='C6E2FF', fill_type="solid")
                            cell.value = None
            self.write_color_counts_and_set_font(workbook)
            workbook.save(excel_path)
            workbook.close()
            
            # 更新UI上的MAP圖顯示
            self.displaySelectedSheet()            
            QMessageBox.information(self, '完成', '初始化所有MAP完成')
            # 初始化完成後啟用Defect Chip summary按鈕
            self.defectChipSummaryBtn.setEnabled(True)
    
    def selectConfirmPath(self):
        # 開啟資料夾對話框選擇確認照片的資料夾
        folder_path = QFileDialog.getExistingDirectory(self, '選擇Lot scanresults')
        if folder_path:
            self.confirmPathLabel.setText(folder_path)
            self.confirmPathLabel.show()

            # 檢查資料夾名稱是否符合規範
            for folder_name in os.listdir(folder_path):
                last_two_digits = folder_name[-2:]
                if not last_two_digits.isdigit() or int(last_two_digits) < 1 or int(last_two_digits) > 25:
                    QMessageBox.warning(self, '警告', '資料名稱不符，File name is not 01~25')
                    break

            self.checkShowRunButton()

    def checkShowRunButton(self):
        # 檢查是否顯示執行按鈕
        if self.pathLabel.text() and self.confirmPathLabel.text():
            self.runSaveBtn.show()
        else:
            self.runSaveBtn.hide()

    def executeSave(self):
        # 執行儲存
        self.executeCore(self.pathLabel.text())

    def outputYieldLossResult(self):
        try:
            if not self.pathLabel.text():
                QMessageBox.warning(self, '警告', '請先選擇MAP圖檔案')
                return

            excel_path = self.pathLabel.text()
            workbook = openpyxl.load_workbook(excel_path, data_only=True)

            # 檢查所有sheet的A欄，收集所有Issue Code Names
            all_issue_code_names = set()
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in range(12, worksheet.max_row + 1):
                    cell = worksheet[f'A{row}']
                    fill_color = self.get_fill_color(cell)
                    if fill_color and (fill_color.red(), fill_color.green(), fill_color.blue()) != (198, 226, 255):
                        if cell.value:
                            all_issue_code_names.add(cell.value)

            if not all_issue_code_names:
                QMessageBox.warning(self, '警告', '沒有找到任何Issue code,無法輸出Yield loss result')
                return

            # 將 all_issue_code_names 轉換為列表並排序
            all_issue_code_names = sorted(list(all_issue_code_names))

            # 讓用戶選擇要排除的Issue Code Name
            excluded_issue_code_names = self.show_total_count_name_dialog(all_issue_code_names)

            # 定義一個函數來獲取Issue Code的簡短格式（前三位數字）
            def get_short_code(code):
                digits = ''.join(filter(str.isdigit, str(code)))
                return digits[:3] if len(digits) >= 3 else digits

            # 將排除的Issue Code Name轉換為簡短格式
            excluded_issue_codes = [get_short_code(code) for code in excluded_issue_code_names]

            # 重新執行顏色計數
            self.write_color_counts_and_set_font(workbook)

            # 創建一個新的工作簿用於存放Yield loss result
            result_workbook = openpyxl.Workbook()
            result_sheet = result_workbook.active
            result_sheet.title = 'Yield loss result'

            # 設置A列和B列的寬度為14
            result_sheet.column_dimensions['A'].width = 14
            result_sheet.column_dimensions['B'].width = 14

            # 創建字體樣式
            font = Font(name='微軟正黑體', size=9, bold=True)

            # 創建對齊方式
            alignment = Alignment(horizontal='center', vertical='center')

            # 寫入表頭並應用格式
            headers = ['Sheet Name', 'Yield loss (%)']
            for col, header in enumerate(headers, start=1):
                cell = result_sheet.cell(row=1, column=col, value=header)
                cell.font = font
                cell.alignment = alignment

            # 創建一個日誌文件來記錄計算過程
            log_file_path = os.path.join(os.path.dirname(excel_path), 'yield_loss_result_by_user_log.txt')
            with open(log_file_path, 'w', encoding='utf-8') as log_file:
                # 將每個sheet的Yield loss寫入結果工作簿，並應用格式
                row_index = 2
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    log_file.write(f"\n--- Sheet: {sheet_name} ---\n")
                    
                    # 計算 Yield loss
                    electrical_fill_count = 0
                    over_kill_count = 0
                    defect_count = 0
                    total_issue_code_count = 0
                    
                    issue_code_counts = {}
                    
                    for row in range(12, worksheet.max_row + 1):
                        cell_a = worksheet[f'A{row}']
                        cell_b = worksheet[f'B{row}']
                        
                        if cell_a.value == 'Electrical_fill':
                            electrical_fill_count = int(cell_b.value) if cell_b.value else 0
                            log_file.write(f"Electrical_fill count: {electrical_fill_count}\n")
                        elif cell_a.value == 'Over_kill':
                            over_kill_count = int(cell_b.value) if cell_b.value else 0
                            log_file.write(f"Over_kill count: {over_kill_count}\n")
                        elif cell_a.value and cell_a.value not in ['Total Defect Count', 'Electrical_fill', 'Over_kill', 'Yield loss (%)']:
                            count = int(cell_b.value) if cell_b.value else 0
                            
                            issue_code_short = get_short_code(cell_a.value)
                            if issue_code_short not in ['151', 'UGL'] and 'Ugly' not in cell_a.value.upper():
                                total_issue_code_count += count  # 所有非 Ugly Die 相關的 Issue codes 都計入總數
                                if issue_code_short not in excluded_issue_codes:
                                    defect_count += count
                                    issue_code_counts[cell_a.value] = count
                                    log_file.write(f"Issue code {cell_a.value}: {count}\n")
                                else:
                                    log_file.write(f"Excluded Issue code {cell_a.value}: {count}\n")
                            else:
                                log_file.write(f"Ignored Issue code {cell_a.value}: {count}\n")

                    yield_loss_denominator = electrical_fill_count + over_kill_count + total_issue_code_count
                    yield_loss = (defect_count / yield_loss_denominator) * 100 if yield_loss_denominator != 0 else 0

                    log_file.write(f"\nTotal defect count: {defect_count}\n")
                    log_file.write(f"Total issue code count (excluding Ugly Die): {total_issue_code_count}\n")
                    log_file.write(f"Yield loss denominator: {yield_loss_denominator}\n")
                    log_file.write(f"Calculated Yield loss: {yield_loss:.2f}%\n")
                    log_file.write("\nDetailed issue code counts:\n")
                    for issue_code, count in issue_code_counts.items():
                        log_file.write(f"{issue_code}: {count}\n")

                    result_sheet[f'A{row_index}'] = sheet_name
                    result_sheet[f'B{row_index}'] = round(yield_loss, 2)
                    
                    # 應用格式到新添加的行
                    for col in ['A', 'B']:
                        cell = result_sheet[f'{col}{row_index}']
                        cell.font = font
                        cell.alignment = alignment
                    
                    row_index += 1

            # 儲存Yield loss result工作簿
            output_folder = os.path.dirname(excel_path)
            output_file = os.path.join(output_folder, 'Yield loss result.xlsx')
            result_workbook.save(output_file)

            QMessageBox.information(self, '完成', f'Yield loss result已輸出\n詳細計算過程已保存至：{log_file_path}')
        except Exception as e:
            QMessageBox.critical(self, '錯誤', f'在計算Yield loss時發生錯誤：{str(e)}')
            print(f"錯誤詳情：{traceback.format_exc()}")

    def show_total_count_name_dialog(self, all_total_count_names):
        dialog = QDialog(self)
        dialog.setWindowTitle("選擇要排除的Total Count Name")
        layout = QVBoxLayout()

        label = QLabel("請選擇要排除的Total Count Name:")
        layout.addWidget(label)

        checkboxes = []
        for name in all_total_count_names:
            # 跳過 'Electrical_fill' 和 'ugly' 相關的 Issue codes，不為其創建複選框
            if name == 'Electrical_fill' or 'ugly' in name.lower() or '151' in name:
                continue
            checkbox = QCheckBox(name)
            checkboxes.append(checkbox)
            layout.addWidget(checkbox)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        excluded_names = []
        if dialog.exec() == QDialog.Accepted:
            excluded_names = [checkbox.text() for checkbox in checkboxes if checkbox.isChecked()]

        return excluded_names

    def executeCore(self, save_path):
        # 執行核心功能
        confirm_folder_path = self.confirmPathLabel.text()
        if os.path.exists(confirm_folder_path):
            workbook = openpyxl.load_workbook(self.pathLabel.text())
            
            # 遍歷每個資料夾
            for folder_name in os.listdir(confirm_folder_path):
                last_two_digits = folder_name[-2:]
                if last_two_digits.isdigit() and 1 <= int(last_two_digits) <= 25:
                    # 找到對應的工作表
                    sheet_name = None
                    for ws_name in workbook.sheetnames:
                        if ws_name[-2:] == last_two_digits:
                            sheet_name = ws_name
                            break

                    if sheet_name:
                        worksheet = workbook[sheet_name]
                        
                        # 收集所有子資料夾中的照片座標
                        xy_points_by_subfolder = {}
                        subfolder_path = os.path.join(confirm_folder_path, folder_name)
                        for subfolder_name in self.folder_color_mapping:
                            subfolder_full_path = os.path.join(subfolder_path, subfolder_name)
                            if os.path.exists(subfolder_full_path) and not ('over' in subfolder_name.lower() and 'kill' in subfolder_name.lower()):
                                xy_points = self.extract_xy_points(subfolder_full_path)
                                xy_points_by_subfolder[subfolder_name] = xy_points
                        
                        # 根據優先順序篩選照片座標
                        final_xy_points = {}
                        for subfolder_name, _ in sorted(self.subfolder_priority.items(), key=lambda x: x[1], reverse=False):
                            if subfolder_name in xy_points_by_subfolder:
                                for x, y in xy_points_by_subfolder[subfolder_name]:
                                    if (x, y) not in final_xy_points:
                                        final_xy_points[(x, y)] = subfolder_name
                        
                        # 將篩選後的結果輸出到Excel
                        for (x, y), subfolder_name in final_xy_points.items():
                            try:
                                x_point = int(x) + 4
                                y_point = int(y) + 2
                                if x_point >= 1 and y_point >= 1:
                                    cell = worksheet.cell(row=y_point, column=x_point)
                                    color = self.folder_color_mapping[subfolder_name]
                                    
                                    # 檢查是否要覆蓋已存在的Issue code
                                    if not self.overwriteCheckBox.isChecked() or not cell.value:
                                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                        cell.value = subfolder_name[:3]
                            except ValueError:
                                pass
                        
                        # 寫入顏色計數並設置字體
                        self.write_color_counts_and_set_font(workbook)

            workbook.save(save_path)
            workbook.close()

            # 顯示當前選擇的工作表的MAP圖
            self.displaySelectedSheet()
            QMessageBox.information(self, '完成', '執行完成')

    def write_color_counts_and_set_font(self, workbook):
        #(Don't Delete) log_file_path = os.path.join(os.path.dirname(self.pathLabel.text()), 'Yield_loss_result_normal_log.txt')
        #(Don't Delete) with open(log_file_path, 'w', encoding='utf-8') as log_file:
        for worksheet in workbook.worksheets:
            #(Don't Delete) log_file.write(f"\n--- Sheet: {worksheet.title} ---\n")
            
            #(Don't Delete) Clear existing data and formatting starting from A12
            for row in range(12, worksheet.max_row + 1):
                worksheet[f'A{row}'].value = None
                worksheet[f'A{row}'].fill = openpyxl.styles.PatternFill(fill_type=None)
                worksheet[f'B{row}'].value = None
                worksheet[f'B{row}'].fill = openpyxl.styles.PatternFill(fill_type=None)

            # Dictionary to store issue code counts
            issue_code_counts = {}

            # Count occurrences of issue codes
            for row in worksheet.iter_rows():
                for cell in row:
                    cell_value = str(cell.value)
                    fill_color = self.get_fill_color(cell)
                    if fill_color:
                        if fill_color.red() == 198 and fill_color.green() == 226 and fill_color.blue() == 255:
                            issue_code_counts['Over_kill'] = issue_code_counts.get('Over_kill', 0) + 1
                        elif fill_color.red() == 0 and fill_color.green() == 0 and fill_color.blue() == 0:
                            issue_code_counts['Electrical_fill'] = issue_code_counts.get('Electrical_fill', 0) + 1
                        elif len(cell_value) == 3 and cell_value.isdigit():
                            issue_code_counts[cell_value] = issue_code_counts.get(cell_value, 0) + 1

            # Write issue code counts to Excel and calculate Yield loss
            row_counter = 12
            electrical_fill_count = issue_code_counts.get('Electrical_fill', 0)
            over_kill_count = issue_code_counts.get('Over_kill', 0)
            defect_count = 0
            total_issue_code_count = 0

            #(Don't Delete) log_file.write(f"Electrical_fill count: {electrical_fill_count}\n")
            #(Don't Delete) log_file.write(f"Over_kill count: {over_kill_count}\n")

            for issue_code, count in issue_code_counts.items():
                color_cell = f'A{row_counter}'
                count_cell = f'B{row_counter}'

                if issue_code == 'Over_kill':
                    issue_name = 'Over_kill'
                    issue_color = 'C6E2FF'
                elif issue_code == 'Electrical_fill':
                    issue_name = 'Electrical_fill'
                    issue_color = '000000'
                else:
                    # 首先嘗試使用 folder_color_mapping
                    issue_name = next((name for name, hex in self.folder_color_mapping.items() if name.startswith(issue_code)), '')
                    issue_color = self.folder_color_mapping.get(issue_name, None)

                    # 如果在 folder_color_mapping 中找不到，嘗試使用 No_folder_color_mapping
                    if not issue_name:
                        issue_name = self.No_folder_color_mapping.get(issue_code, issue_code)
                        # 如果在 No_folder_color_mapping 中找到了新名稱，使用原始的 issue_code 顏色
                        if issue_name != issue_code:
                            for row in worksheet.iter_rows():
                                for cell in row:
                                    if str(cell.value) == issue_code:
                                        issue_color = self.get_fill_color(cell)
                                        if issue_color:
                                            issue_color = f"{issue_color.red():02X}{issue_color.green():02X}{issue_color.blue():02X}"
                                            break
                                if issue_color:
                                    break

                worksheet[color_cell].fill = openpyxl.styles.PatternFill(start_color=issue_color, end_color=issue_color, fill_type="solid")
                worksheet[color_cell].value = issue_name
                worksheet[count_cell].value = count

                arial_font = openpyxl.styles.Font(name='微軟正黑體', size=9)
                worksheet[color_cell].font = arial_font
                worksheet[count_cell].font = arial_font

                #(Don't Delete) log_file.write(f"Issue code {issue_name}: {count}\n")

                if issue_code not in ['Over_kill', 'Electrical_fill']:
                    if issue_code != '151' and 'Ugly' not in issue_name.upper():
                        total_issue_code_count += count
                        defect_count += count

                row_counter += 1

            # Calculate Yield loss
            yield_loss_denominator = electrical_fill_count + over_kill_count + total_issue_code_count
            yield_loss = (defect_count / yield_loss_denominator) * 100 if yield_loss_denominator != 0 else 0

            #(Don't Delete) log_file.write(f"\nTotal defect count: {defect_count}\n")
            #(Don't Delete) log_file.write(f"Total issue code count (excluding Ugly Die): {total_issue_code_count}\n")
            #(Don't Delete) log_file.write(f"Yield loss denominator: {yield_loss_denominator}\n")
            #(Don't Delete) log_file.write(f"Calculated Yield loss: {yield_loss:.2f}%\n")

            yield_loss_cell = f'A{row_counter}'
            yield_loss_value_cell = f'B{row_counter}'
            worksheet[yield_loss_cell].value = 'Yield loss (%)'
            worksheet[yield_loss_value_cell].value = round(yield_loss, 2)

            arial_font = openpyxl.styles.Font(name='微軟正黑體', size=9)
            worksheet[yield_loss_cell].font = arial_font
            worksheet[yield_loss_value_cell].font = arial_font

    def extract_xy_points(self, folder_path):
        # 從照片檔名提取XY座標
        xy_points = []
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                # 現有的模式
                if "-" in filename:
                    match = re.search(r'-\d+_(\d+)_(\d+)_', filename)
                else:
                    match = re.search(r'[^_]*_[^_]*_([0-9]+)_([0-9]+)_', filename)

                # 新的模式，用於像是 "KK4QR07_2_8_Un-reviewed_1.jpg" 這樣的檔名
                if not match:
                    match = re.search(r'[^_]*_([0-9]+)_([0-9]+)_', filename)

                if match:
                    try:
                        x_point = int(match.group(1))
                        y_point = int(match.group(2))
                    except ValueError:
                        raise ValueError(f"檔名中的 x_point 或 y_point 無效: {filename}")

                    xy_points.append((x_point, y_point))
                else:
                    raise ValueError(f"檔名不符合模式: {filename}")

        return xy_points

    def get_fill_color(self, cell):
        # 獲取儲存格的填充顏色
        fill = cell.fill
        if fill.fill_type == "solid" and hasattr(fill.start_color, 'rgb'):
            try:
                return QColor.fromRgb(*[int(fill.start_color.rgb[i:i+2], 16) for i in (2, 4, 6)])
            except ValueError:
                # 處理轉換錯誤
                return None
        else:
            return None

    def displaySelectedSheet(self):
        # 顯示選擇的Sheet的MAP圖
        sheet_name = self.sheetComboBox.currentText()
        self.displayMap(self.pathLabel.text(), sheet_name)

    def displayMap(self, excel_path, sheet_name):
        # 顯示地圖
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            worksheet = workbook[sheet_name]
        except Exception as e:
            QMessageBox.critical(self, '錯誤', f'發生錯誤: {e}')
            return

        # 找到有值的範圍
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        start_col = 3  # 從C欄開始

        # 找到最後一欄有值的位置
        end_col = max_col
        while end_col >= start_col and all(worksheet.cell(row, end_col).value is None for row in range(1, max_row + 1)):
            end_col -= 1

        # 在找到的最後一欄有值的位置後再往右增加一欄
        end_col += 1

        if end_col >= start_col:
            # 設定單元格大小
            cell_width = 8
            cell_height = 6
            
            # 計算圖片大小
            image_width = (end_col - start_col) * cell_width
            image_height = max_row * cell_height
            
            # 創建一張新的圖片
            image = QImage(image_width, image_height, QImage.Format_RGB32)
            image.fill(Qt.white)
            painter = QPainter(image)
            
            # 繪製表格，包括儲存格的底色
            for row in range(max_row):
                for col in range(start_col, end_col):
                    cell = worksheet.cell(row + 1, col)
                    cell_value = cell.value
                    cell_rect = QRect((col - start_col) * cell_width, row * cell_height, cell_width, cell_height)
                    
                    # 獲取儲存格的填充顏色
                    fill_color = self.get_fill_color(cell)
                    if fill_color:
                        painter.fillRect(cell_rect, fill_color)  # 使用填充顏色填充矩形
                    
                    painter.drawRect(cell_rect)
                    if cell_value:
                        painter.drawText(cell_rect, Qt.AlignCenter, str(cell_value))
            
            painter.end()
            
            # 顯示圖片
            pixmap = QPixmap.fromImage(image)
            self.imageLabel.setPixmap(pixmap)
            self.imageLabel.setScaledContents(True)  # 保持圖片的原始尺寸
            self.imageLabel.setAlignment(Qt.AlignCenter)
        else:
            self.imageLabel.clear()
            self.imageLabel.setText('沒有資料')

        # 保存修改後的工作簿
        workbook.save(excel_path)
        workbook.close()

    def performDefectChipSummary(self):
        try:
            if not self.pathLabel.text():
                QMessageBox.warning(self, '警告', '請先選擇MAP圖檔案')
                return
            
            if not self.defectChipSummaryBtn.isEnabled():
                QMessageBox.warning(self, '警告', '請先使用initializeAllMapsBtn按鈕')
                return
            
            dialog = QFileDialog()
            dialog.setFileMode(QFileDialog.Directory)  
            dialog.setOption(QFileDialog.DontUseNativeDialog, True)
            dialog.setOption(QFileDialog.ShowDirsOnly, False) 
            dialog.setViewMode(QFileDialog.Detail) 
            dialog.resize(1000, 600)
            dialog.setStyleSheet("""
                QFileDialog, QLabel, QPushButton, QComboBox, QLineEdit, QListView, QTreeView {
                    background-color: white;
                    color: black;
                }
                QFileDialog QListView::item:selected, QFileDialog QTreeView::item:selected {
                    background-color: #0078D7;
                    color: white;
                }
            """)
            dialog.setWindowTitle('選擇要進行累計Defct count by chip的資料夾路徑,資料夾內必須確保Map Excel都是相同的Group!!')
            
            if dialog.exec_() == QDialog.Accepted:
                folder_path = dialog.selectedFiles()[0]
            else:
                return
            
            # 讀取選擇的MAP圖檔案
            map_file_path = self.pathLabel.text()
            map_workbook = openpyxl.load_workbook(map_file_path)
            
            # 選擇第一個工作表作為Format
            sheet_name = map_workbook.sheetnames[0]
            
            # 刪除選擇的Sheet以外的其他Sheet
            selected_sheet = map_workbook[sheet_name]
            for sheet_name in map_workbook.sheetnames:
                if sheet_name != selected_sheet.title:
                    map_workbook.remove(map_workbook[sheet_name])
            
            # 將選擇的Sheet改名為"ChipSum01"
            selected_sheet.title = "ChipSum01"
            
            # 將修改後的MAP圖檔案另存為"ChipSum result.xlsx"
            chipsum_result_path = os.path.join(folder_path, 'ChipSum result.xlsx')
            map_workbook.save(chipsum_result_path)
            map_workbook.close()
            
            #關閉按鈕使用權限
            self.outputYieldLossResultBtn.hide()
            self.outputDefectDieBtn.hide()
            self.overwriteCheckBox.hide()
            self.confirmPathBtn.hide()
            self.sheetComboBox.hide()
            self.initializeAllMapsBtn.hide()
            self.defectChipSummaryBtn.hide()
            self.importBtn.hide()

            # 删除原始的MAP圖檔案
            os.remove(map_file_path)
            
            # 更新UI導入MAP圖的路徑
            self.pathLabel.setText(chipsum_result_path)
            
            # 更新UI上的MAP圖顯示
            self.displayMap(chipsum_result_path, 'ChipSum01')
            
            # 讀取使用者選擇資料夾路徑內的所有.xlsx檔案
            cell_counts = {}
            color_names = {}
            for file_name in os.listdir(folder_path):
                if file_name.endswith('.xlsx') and file_name != 'ChipSum result.xlsx':
                    file_path = os.path.join(folder_path, file_name)
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                    
                    for sheet in workbook.worksheets:
                        # 獲取A欄的顏色名稱
                        for row in range(1, sheet.max_row + 1):
                            cell = sheet.cell(row=row, column=1)
                            fill_color = self.get_fill_color(cell)
                            if fill_color:
                                fill_color_hex = f"{fill_color.red():02X}{fill_color.green():02X}{fill_color.blue():02X}"
                                if fill_color_hex not in color_names:
                                    color_names[fill_color_hex] = cell.value
                        
                        # 累計C欄以後的儲存格
                        for row in sheet.iter_rows(min_col=3):
                            for cell in row:
                                cell_value = str(cell.value)
                                fill_color = self.get_fill_color(cell)
                                if fill_color and len(cell_value) >= 3 and cell_value.isdigit():
                                    cell_coordinate = cell.coordinate
                                    cell_fill_color = f"{fill_color.red():02X}{fill_color.green():02X}{fill_color.blue():02X}"
                                    
                                    if cell_coordinate not in cell_counts:
                                        cell_counts[cell_coordinate] = {'count': 0, 'colors': {}}
                                    
                                    cell_counts[cell_coordinate]['count'] += 1
                                    if cell_fill_color not in cell_counts[cell_coordinate]['colors']:
                                        cell_counts[cell_coordinate]['colors'][cell_fill_color] = {'count': 0, 'name': color_names.get(cell_fill_color, '')}
                                    cell_counts[cell_coordinate]['colors'][cell_fill_color]['count'] += 1
                    
                    workbook.close()
            
            # 將統計結果登記到"ChipSum01"工作表內
            chipsum_workbook = openpyxl.load_workbook(chipsum_result_path)
            chipsum_sheet = chipsum_workbook['ChipSum01']
            for cell_coordinate, data in cell_counts.items():
                chipsum_sheet[cell_coordinate].value = data['count']
                
                # 找到出現次數最多的顏色
                max_color = max(data['colors'], key=lambda x: data['colors'][x]['count'])
                chipsum_sheet[cell_coordinate].fill = PatternFill(start_color=max_color, end_color=max_color, fill_type='solid')
                
                # 建立註解內容
                comment_text = []
                sorted_colors = sorted(data['colors'].values(), key=lambda x: x['count'], reverse=True)
                for color_data in sorted_colors:
                    if color_data['name']:
                        comment_text.append(f"{color_data['name']}: {color_data['count']}次")
                
                # 如果有註解內容,就在儲存格新增註解
                if comment_text:
                    comment = openpyxl.comments.Comment('\n'.join(comment_text), "ChipSum")
                    chipsum_sheet[cell_coordinate].comment = comment
            
            # 創建微軟正黑體字體,大小為9的Font對象
            font = openpyxl.styles.Font(name='微軟正黑體', size=9)
            
            # 將顏色名稱寫入A欄
            row_counter = 12
            for color_hex, color_name in color_names.items():
                color_cell = f'A{row_counter}'
                chipsum_sheet[color_cell].value = color_name
                chipsum_sheet[color_cell].fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                chipsum_sheet[color_cell].font = font  # 設置字體
                row_counter += 1

            for row in range(5, chipsum_sheet.max_row + 1):
                if 5 <= row <= 7 or 9 <= row <= 10 or row >= 12:
                    chipsum_sheet[f'B{row}'].value = None
            
            # 根據A欄的顏色名稱,複製工作表並進行分類
            for row in range(12, chipsum_sheet.max_row + 1):
                color_name_cell = chipsum_sheet[f'A{row}']
                color_name = color_name_cell.value
                fill_color = self.get_fill_color(color_name_cell)
                
                if color_name and fill_color:
                    # 複製ChipSum01工作表
                    copied_sheet = chipsum_workbook.copy_worksheet(chipsum_sheet)
                    
                    # 將工作表名稱中的特殊符號替換為空格,如果有相連兩個空格,則維持一個
                    sheet_name = re.sub(r'[^a-zA-Z0-9\s]', ' ', str(color_name))
                    sheet_name = re.sub(r'\s+', ' ', sheet_name)
                    copied_sheet.title = sheet_name
                    
                    # 將C欄以後顏色不符的儲存格顏色轉換為C6E2FF,並刪除儲存格值
                    fill_color_hex = f"{fill_color.red():02X}{fill_color.green():02X}{fill_color.blue():02X}"
                    for row in copied_sheet.iter_rows(min_col=3):
                        for cell in row:
                            cell_fill_color = self.get_fill_color(cell)
                            if cell_fill_color:
                                cell_fill_color_hex = f"{cell_fill_color.red():02X}{cell_fill_color.green():02X}{cell_fill_color.blue():02X}"
                                if cell_fill_color_hex != fill_color_hex:
                                    cell.fill = PatternFill(start_color='C6E2FF', end_color='C6E2FF', fill_type='solid')
                                    cell.value = None
                                    if cell.comment:  # 檢查是否有註解
                                        cell.comment = None  # 刪除註解內容
                    
                    # 清空A欄,只保留A12儲存格的顏色和值,並將A12欄以後到A99儲存格設定為無框線
                    for row in range(12, 100):
                        cell = copied_sheet[f'A{row}']
                        if row == 12:
                            cell.value = color_name
                            cell.fill = PatternFill(start_color=fill_color_hex, end_color=fill_color_hex, fill_type='solid')
                            cell.font = font
                            cell.border = openpyxl.styles.Border()
                        else:
                            cell.value = None
                            cell.fill = PatternFill(fill_type=None)
                            cell.border = openpyxl.styles.Border()

            # 創建一個新的工作表 'Lots list'
            lots_list_sheet = chipsum_workbook.create_sheet('Lots list')

            # 設定 'Lots list' 工作表的A欄寬為50
            lots_list_sheet.column_dimensions['A'].width = 50

            # 創建微軟正黑體字體,大小為9的Font對象
            font = openpyxl.styles.Font(name='微軟正黑體', size=9)

            # 將讀取到的所有 .xlsx 檔案名稱寫入 'Lots list' 工作表的A欄
            row_index = 1
            for file_name in os.listdir(folder_path):
                if file_name.endswith('.xlsx') and file_name != 'ChipSum result.xlsx':
                    lots_list_sheet.cell(row=row_index, column=1, value=file_name)
                    lots_list_sheet.cell(row=row_index, column=1).font = font
                    row_index += 1

            # 保存"ChipSum result.xlsx"
            chipsum_workbook.save(chipsum_result_path)
            chipsum_workbook.close()
            
            # 更新UI上的MAP圖顯示
            self.displayMap(chipsum_result_path, 'ChipSum01')
            
            QMessageBox.information(self, '完成', '累計Defct count by chip完成')

            # 打開包含"ChipSum result.xlsx"的資料夾
            os.startfile(folder_path)

        except Exception as e:
            # 記錄錯誤日誌
            error_message = f"執行過程中發生錯誤: {str(e)}"
            print(error_message)
            
            # 顯示錯誤提示框
            QMessageBox.critical(self, '錯誤', error_message)

    def save_log(self):
        try:
            hostname = socket.gethostname()
            match = re.search(r'^(.+)', hostname)
            username = match.group(1) if match else 'Unknown'

            current_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_folder = r'M:\QA_Program_Raw_Data\Log History'
            archive_path = os.path.join(log_folder, 'Map confirm.7z')
            log_filename = f'{username}.txt'
            new_log_message = f"{current_datetime} {username} Open\n"
            os.makedirs(log_folder, exist_ok=True)

            if not os.path.exists(archive_path):
                with py7zr.SevenZipFile(archive_path, mode='w', password='@Joe11111111') as archive:
                    archive.writestr(new_log_message, f'Map confirm/{log_filename}')
            else:
                log_content = ""
                files_to_keep = []

                with py7zr.SevenZipFile(archive_path, mode='r', password='@Joe11111111') as archive:
                    for filename, bio in archive.read().items():
                        if filename == f'Map confirm/{log_filename}':
                            log_content = bio.read().decode('utf-8')
                        else:
                            files_to_keep.append((filename, bio.read()))

                if new_log_message not in log_content:
                    log_content += new_log_message

                with tempfile.NamedTemporaryFile(delete=False, suffix='.7z') as temp_file:
                    temp_archive_path = temp_file.name

                with py7zr.SevenZipFile(temp_archive_path, mode='w', password='@Joe11111111') as archive:
                    archive.writestr(log_content.encode('utf-8'), f'Map confirm/{log_filename}')
                    for filename, content in files_to_keep:
                        archive.writestr(content, filename)

                shutil.move(temp_archive_path, archive_path)

        except Exception as e:
            print(f"寫入log時發生錯誤: {e}")

    def check_version(self):
        try:
            app_folder = r"M:\QA_Program_Raw_Data\Apps"
            exe_files = [f for f in os.listdir(app_folder) if f.startswith("Map confirm_V") and f.endswith(".exe")]

            if not exe_files:
                QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                sys.exit(1)

            # 修改版本號提取邏輯，只取主版本號
            latest_version = max(int(re.search(r'_V(\d+)', f).group(1)) for f in exe_files)

            # 修改當前版本號提取邏輯，只取主版本號
            current_version_match = re.search(r'_V(\d+)', os.path.basename(sys.executable))
            if current_version_match:
                current_version = int(current_version_match.group(1))
            else:
                current_version = 23

            if current_version < latest_version:
                QMessageBox.information(self, '請更新至最新版本', '請更新至最新版本')
                os.startfile(app_folder)  # 開啟指定的資料夾
                sys.exit(0)

            hostname = socket.gethostname()
            match = re.search(r'^(.+)', hostname)
            if match:
                username = match.group(1)
                if username == "A000000":
                    QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                    sys.exit(1)
            else:
                QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
                sys.exit(1)

        except FileNotFoundError:
            QMessageBox.warning(self, '未獲取啟動權限', '未獲取啟動權限, 請申請M:\QA_Program_Raw_Data權限, 並聯絡#1082 Racky')
            sys.exit(1)

    def on_checkbox_state_changed(self, state, code):
        if state == Qt.Checked:
            self.selected_issue_codes.append(code)
        else:
            self.selected_issue_codes.remove(code)

    def show_issue_code_dialog(self, all_issue_codes, default_excluded_codes):
        # 如果沒有 Issue Code，直接返回空列表
        if not all_issue_codes:
            return []

        dialog = QDialog(self)
        dialog.setWindowTitle("選擇要排除的Issue Code")
        layout = QVBoxLayout()

        label = QLabel("請選擇要排除的 Issue Code：")
        layout.addWidget(label)

        checkboxes = []
        for issue_code in all_issue_codes:
            # 檢查 Issue Code 的前三位是否為數字
            if issue_code[:3].isdigit():
                checkbox = QCheckBox(issue_code)
                if any(issue_code.startswith(code) for code in default_excluded_codes):
                    checkbox.setChecked(True)
                checkboxes.append(checkbox)
                layout.addWidget(checkbox)

        # 如果沒有有效的 Issue Code 可選，直接返回空列表
        if not checkboxes:
            return []

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        excluded_issue_codes = []
        if dialog.exec() == QDialog.Accepted:
            excluded_issue_codes = [checkbox.text() for checkbox in checkboxes if checkbox.isChecked()]

        return excluded_issue_codes

    def outputDefectDieCoordinates(self):
        if not self.pathLabel.text():
            QMessageBox.warning(self, '警告', '請先選擇MAP圖檔案')
            return

        excel_path = self.pathLabel.text()
        workbook = openpyxl.load_workbook(excel_path, data_only=True)

        # 檢查所有sheet的A欄，收集所有Issue Code Names
        all_issue_code_names = set()
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in range(12, worksheet.max_row + 1):
                cell = worksheet[f'A{row}']
                if cell.value:
                    all_issue_code_names.add(cell.value)

        # 如果沒有找到任何Issue code，顯示警告並返回
        if not all_issue_code_names:
            QMessageBox.warning(self, '警告', '沒有找到任何Issue code，無法輸出Defect die座標')
            return

        # 將 all_issue_code_names 轉換為列表並排序
        all_issue_code_names = sorted(list(all_issue_code_names))

        # 預設要勾選的Issue Code
        default_excluded_codes = ['151', '000', '999', '186']

        # 讓用戶選擇要排除的Issue Code Name
        excluded_issue_code_names = self.show_issue_code_dialog(all_issue_code_names, default_excluded_codes)

        # 定義一個函數來獲取Issue Code的簡短格式（前三位數字）
        def get_short_code(code):
            digits = ''.join(filter(str.isdigit, str(code)))
            return digits[:3] if len(digits) >= 3 else digits

        # 將排除的Issue Code Name轉換為簡短格式
        excluded_issue_codes = [get_short_code(code) for code in excluded_issue_code_names]

        output_files_created = False
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            defect_die_coordinates = []

            for row in range(1, worksheet.max_row + 1):
                for col in range(3, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = str(cell.value)
                    if len(cell_value) == 3 and cell_value.isdigit() and self.get_fill_color(cell):
                        defect_die_issue_code_number = cell_value

                        if defect_die_issue_code_number not in excluded_issue_codes:
                            x = col - 4
                            y = row - 2
                            defect_die_coordinates.append(f"{x},{y},{defect_die_issue_code_number}")

            if defect_die_coordinates:
                output_folder = os.path.dirname(excel_path)
                output_file = os.path.join(output_folder, f"{sheet_name}.txt")
                with open(output_file, 'w') as file:
                    file.write('\n'.join(defect_die_coordinates))
                output_files_created = True

        if output_files_created:
            QMessageBox.information(self, '完成', 'Defect die座標已輸出')
        else:
            QMessageBox.warning(self, '警告', '沒有符合條件的Defect die座標可輸出')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    font = QFont("微軟正黑體", 9)
    font.setBold(True)
    app.setFont(font)
    app.setWindowIcon(QIcon(icon_path))
    qtmodern.styles.dark(app)

    window = ImageClassifier()
    window.check_version()
    win = qtmodern.windows.ModernWindow(window)
    win.show()

    # 獲取螢幕的寬度和高度
    screen = app.primaryScreen()
    screen_width = screen.geometry().width()
    screen_height = screen.geometry().height()

    # 計算視窗的左上角座標，使其在螢幕中央
    window_width = win.frameGeometry().width()
    window_height = win.frameGeometry().height()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # 設定視窗位置
    win.move(x, y)
    sys.exit(app.exec_())  